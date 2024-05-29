import xlrd
from openpyxl import Workbook, load_workbook

def convert_xls_to_xlsx(xls_filename, xlsx_filename):
    # Open the .xls file
    xls_workbook = xlrd.open_workbook(xls_filename)
    
    # Create a new .xlsx file
    xlsx_workbook = Workbook()
    
    # Remove the default sheet created by openpyxl
    xlsx_workbook.remove(xlsx_workbook.active)
    
    # Copy each sheet from the .xls file to the .xlsx file
    for sheet_index in range(xls_workbook.nsheets):
        xls_sheet = xls_workbook.sheet_by_index(sheet_index)
        xlsx_sheet = xlsx_workbook.create_sheet(title=xls_sheet.name)
        
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                xlsx_sheet.cell(row=row+1, column=col+1).value = xls_sheet.cell_value(row, col)
    
    # Save the .xlsx file
    xlsx_workbook.save(xlsx_filename)

def read_xlsx(filename):
    # Open the .xlsx file using openpyxl
    wb = load_workbook(filename)
    
    # Loop through all sheets
    for sheet in wb:
        print(f"Reading sheet: {sheet.title}")
        # Print the content of the first cell in each sheet
        print(f"Content of the first cell in sheet {sheet.title}: {sheet['A1'].value}")

# Convert the first .xls file
convert_xls_to_xlsx('L070n087_SET_Export.xls', 'L070n087_SET_Export.xlsx')

# Convert the second .xls file
convert_xls_to_xlsx('L070n087_EDE.xls', 'L070n087_EDE.xlsx')

# Read the converted .xlsx files
read_xlsx('L070n087_SET_Export.xlsx')
read_xlsx('L070n087_EDE.xlsx')
