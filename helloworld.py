import xlrd
import openpyxl

ede = op.load_workbook('L070n087_EDE.xls')


set_export = op.load_workbook('L070n087_SET_Export')


# print(set_export.sheetnames)


# Open the .xls file using xlrd
workbook = xlrd.open_workbook('L070n087_SET_Export.xls')

# Access the first sheet by index
sheet = workbook.sheet_by_index(0)

# Print the content of the first cell in the first row

print(sheet.cell_value(0, 0))
print(sheet.row_values(1))