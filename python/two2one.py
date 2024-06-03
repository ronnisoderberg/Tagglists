
from openpyxl import Workbook, load_workbook
from  xls2xlsx import convertXlsToXlsx, read_xlsx

convert_xls_to_xlsx('L070n087_SET_Export.xls', 'L070n087_SET_Export.xlsx')

read_xlsx('L070n087_SET_Export.xlsx')

scadaLista = Workbook()

ws1 = scadaLista.create_sheet('Tagglista', 0)
ws2 = scadaLista.create_sheet('Larm')

scadaLista.save('scadaLista.xlsx')

convertXlsToXlsx('L070n087_SET_Export.xls', 'L070n087_SET_Export.xlsx')
convertXlsToXlsx('L070n087_EDE.xls', 'L070n087_EDE.xlsx')


read_xlsx('L070n087_SET_Export.xlsx')
read_xlsx('L070n087_EDE.xlsx')

scadaList_xlsx = Workbook()
scadaList_xlsx.remove(scadaList_xlsx.active)
wsLarm = scadaList_xlsx.create_sheet(title='Larm')
wsVariabler = scadaList_xlsx.create_sheet(title='Variabler')


scadaList_xlsx.save('scadaList.xlsx')